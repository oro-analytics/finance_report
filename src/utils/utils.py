import re
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.utils.get_headers import pl_header, secured_rev_header

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def extract_profit_center_data(file_path: Path, target_profit_center: str):
    try:
        # Извлекаем месяц из названия файла: PL_02 2025.xlsx -> 02
        match = re.search(r"PL_(\d{2})", file_path.stem)
        month = int(match.group(1)) if match else None
        year = int(file_path.parent.name)

        df = pd.read_excel(file_path, sheet_name="pl_projects")

        # Отбрасываем строки без Profit Center
        df_cleaned = df[df[pl_header(year, month)["Profit Center"]].notna()]

        # в разных годах заполнено по разному
        target_pc = secured_rev_header(year, month)[target_profit_center]

        # Фильтрация по заданному Profit Center и непустой колонке "Реализация без НДС"
        filtered_df = df_cleaned[
            (df_cleaned[pl_header(year, month)["Profit Center"]] == target_pc) &
            (df_cleaned["Реализация без НДС"].notna())
        ]

        total_revenue = filtered_df["Реализация без НДС"].sum()
        total_direct_cost = filtered_df[pl_header(year, month)["Total Direct Costs"]].sum()
        total_operating_cost = filtered_df[pl_header(year, month)["Total Operating Costs"]].sum()

        filtered_df_out = filtered_df.copy()
        filtered_df_out["Год"] = year
        filtered_df_out["Месяц"] = month
        filtered_df_out["Источник файла"] = file_path.name

        # сохраняем название профит-центра из справочника в отдельной колонке
        profit_center_column = pl_header(year, month)["Profit Center"]
        if profit_center_column in filtered_df_out.columns:
            filtered_df_out["Название Profit Center"] = filtered_df_out[profit_center_column]

        summary_row = {
            "Файл": file_path.name,
            "Год": year,
            "Месяц": month,
            "Profit Center": target_pc,
            "Сумма 'Реализация без НДС'": total_revenue,
            "Сумма 'Total Direct Costs'": total_direct_cost,
            "Сумма 'Total Operating Costs'": total_operating_cost,
            "Operation Profit": total_revenue - total_direct_cost - total_operating_cost
        }

        return summary_row, filtered_df_out
    except Exception as e:
        return {
            "Файл": file_path.name,
            "Год": file_path.parent.name,
            "Месяц": None,
            "Ошибка": str(e)
        }, pd.DataFrame()


def extract_df_with_combined_header(df2):
    """
        Извлекает DataFrame начиная с start_row,
        удаляет пустые строки и колонки,
        объединяет первые две строки в заголовок.
        """

    # Получаем первые две строки
    header_1 = df2.iloc[0].astype(str).str.strip()
    header_2 = df2.iloc[1].astype(str).str.strip()

    # безопасное преобразование заголовка
    def _clean_header(value):
        if pd.isna(value):
            return ""
        return str(value).strip()

    # Объединяем заголовки: если в первой строке пусто, берём из второй
    combined_header = []
    for h1, h2 in zip(header_1, header_2):
        h1_clean = _clean_header(h1)
        h2_clean = _clean_header(h2)
        combined_header.append(h1_clean if h1_clean and h1_clean.lower() != 'nan' else h2_clean)

    df2.columns = combined_header
    df2 = df2.iloc[2:]  # удаляем строки, которые пошли в заголовки
    df2 = df2.reset_index(drop=True)
    df2 = df2[df2['Profit Center'].notna()]
    return df2


def extract_x_charge_data(file_path: Path, target_profit_center: str):
    try:
        # Извлекаем месяц из названия файла: PL_02 2025.xlsx -> 02
        match = re.search(r"Secured Rev_Profit centers_(\d{2})", file_path.stem)
        month = int(match.group(1)) if match else None
        year = int(file_path.parent.name)

        #
        df_full = pd.read_excel(file_path, sheet_name="Secured Rev - Profit centers", header=None)

        # в разных годах заполнено по разному
        target_pc = secured_rev_header(year, month)[target_profit_center]

        # Пример: первый DataFrame — строки 0 по 9, второй — строки 11 по 20
        df1 = df_full.iloc[0:18].dropna(how='all')  # убираем пустые строки
        df1.columns = df1.iloc[0]  # если первая строка — заголовки
        df1 = df1[1:]
        df1 = df1.dropna(axis=1, how='all')
        row_with_profit_center = df1[df1['Profit center'] == target_pc]

        # Читаем вторую таблицу
        # --- Находим строку начала X-charge автоматически ---
        xcharge_start_row = None
        for i, row in df_full.iterrows():
            if re.search(r"X.?charge", str(row[0]), flags=re.IGNORECASE):
                xcharge_start_row = i
                break

        if xcharge_start_row is None:
            raise ValueError("❌ Не найдена строка начала X-charge. Проверьте содержимое файла.")

        # --- Отрезаем начиная с найденной строки ---
        xcharge_table_start_row = xcharge_start_row + 2
        df2 = df_full.iloc[xcharge_table_start_row:].dropna(axis=0, how='all').dropna(axis=1, how='all')

        # Обрабатываем таблицу как раньше
        df2 = extract_df_with_combined_header(df2)

        df2_giver = df2[df2['Profit Center'].str.contains(target_pc, na=False)]
        cols = [c for c in df2.columns if target_pc in c]
        if cols:
            df2_taker = df2[df2[cols].gt(0).any(axis=1)]
        else:
            df2_taker = pd.DataFrame()
            print(f"\t Для {target_profit_center} НЕ найдены X-charge, которые надо ДОБАВИТЬ")

        if not df2_giver.empty:
            print(f"\t Для {target_profit_center} найдены X-charge, которые надо ОТДАТЬ:")
            print(df2_giver)

        if not df2_taker.empty:
            print(f"### \t Для {target_profit_center} найдены X-charge, которые надо ДОБАВИТЬ:")
            print(df2_taker[['Profit Center', 'Компания', 'Номер контракта',
                             'Сумма контракта без НДС', 'Дата начала контракта',
                             'Дата завершения контракта'] + cols
                             ].T)

        row_with_profit_center = (
            row_with_profit_center.copy().assign(
                Файл=file_path.name,
                Год=year,
                Месяц=month
            )
        ).reset_index(drop=True)

        return row_with_profit_center, df2_giver, df2_taker
    except Exception as e:
        print(f"ошибка: {e}")
        return pd.DataFrame({
            "Файл": file_path.name,
            "Год": file_path.parent.name,
            "Месяц": None,
            "Ошибка": str(e)
        }), pd.DataFrame(), pd.DataFrame()


def process_all_pl_files(base_dir: str, years: list, target_profit_center: str):
    results = []
    monthly_projects: dict[str, pd.DataFrame] = {}

    for year in years:
        year_path = Path(base_dir) / str(year)
        if not year_path.exists():
            continue

        for file in sorted(year_path.glob("PL_*.xlsx")):
            print(f"Working with {file}")
            summary, filtered_df = extract_profit_center_data(file, target_profit_center)
            results.append(summary)

            if not filtered_df.empty and summary.get("Месяц"):
                month_key = f"{int(summary['Месяц']):02d} {year}"
                monthly_projects[month_key] = filtered_df

    return pd.DataFrame(results), monthly_projects


def process_all_x_charge_files(
    base_dir: str,
    years: list[int],
    target_profit_center: str
) -> tuple[dict, dict, dict]:
    """
        Обрабатывает все файлы вида Secured Rev_Profit centers_*.xlsx.
        Возвращает три словаря:
          results_dict[month] = result_df
          giver_dict[month]   = giver_df
          taker_dict[month]   = taker_df
        где month берётся из * в имени файла.
        """

    results_dict, giver_dict, taker_dict = {}, {}, {}

    for year in years:
        year_path = Path(base_dir) / str(year)
        if not year_path.exists():
            continue

        for file in sorted(year_path.glob("Secured Rev_Profit centers_*.xlsx")):
            print(f"Working with {file}")

            # достаём часть из имени файла после последнего "_"
            # пример: "Secured Rev_Profit centers_03.xlsx" -> "03"
            m = re.search(r"Secured Rev_Profit centers_(.+)\.xlsx", file.name)
            if not m:
                print(f"⚠️ не удалось вытащить номер месяца из {file.name}")
                continue
            month_key = m.group(1)

            data = extract_x_charge_data(file, target_profit_center)
            if len(data) < 3:
                print(f"⚠️ неожиданный формат данных в {file.name}")
                continue

            result, df2_giver, df2_taker = data[0], data[1], data[2]

            # кладём напрямую в словари
            results_dict[month_key] = result
            giver_dict[month_key] = df2_giver
            taker_dict[month_key] = df2_taker

    return results_dict, giver_dict, taker_dict


def write_monthly_with_highlights(
    dfs_dict: dict[str, pd.DataFrame],
    output_path: str = "monthly_contracts.xlsx",
    id_col: str = "Номер контракта",
    highlight_first: bool = True,   # для самого первого месяца: подсвечивать все как «впервые»
    add_flag_column: bool = True,   # добавить колонку «Новая запись?»
) -> str:
    """
    dfs — словарь {ключ: DataFrame}, где ключ = месяц/период (используется как имя листа).
    output_path — путь к итоговому Excel.
    id_col — имя столбца-идентификатора.
    highlight_first — подсвечивать ли весь первый месяц (всё «впервые»).
    add_flag_column — добавлять ли текстовый флажок «Новая запись?» в таблицу.

    На каждом листе:
      NEW       — запись появилась впервые (строка подсвечивается зелёным)
      MODIFIED  — запись была ранее, но её поля изменились (жёлтым подсвечиваются только изменённые ячейки)
      DELETED   — запись исчезла по сравнению с прошлым месяцем (строка подсвечивается красным)
    """

    def _normalize_for_compare(frame: pd.DataFrame) -> pd.DataFrame:
        return frame.map(lambda value: "" if pd.isna(value) else str(value))

    if not dfs_dict:
        raise ValueError("Словарь dfs_dict пуст.")

    # Отсортируем ключи (если они строки с числами, можно привести к int)
    try:
        ordered_keys = sorted(dfs_dict.keys(), key=lambda x: int(x))
    except ValueError:
        ordered_keys = sorted(dfs_dict.keys())

    # Стиль подсветки
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # ### для MODIFIED

    prev_ids: set[str] = set()
    prev_df: pd.DataFrame | None = None

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for i, key in enumerate(ordered_keys):
            df = dfs_dict[key]
            if id_col not in df.columns:
                # raise KeyError(f'В DataFrame для {key} нет столбца "{id_col}"')
                continue

            df2 = df.copy()
            df2 = df2[~df2[id_col].isna()].copy()
            df2[id_col] = df2[id_col].astype(str).str.strip()
            df2 = df2.drop_duplicates(subset=[id_col], keep="first")

            curr_ids = df2[id_col]
            curr_ids_set = set(curr_ids)

            changed_ids: set[str] = set()
            deleted_rows_df = pd.DataFrame()
            changed_cells: dict[str, set[str]] = {}  # ### id -> набор изменённых колонок

            if i == 0 and highlight_first:
                # В первый месяц всё считаем "новым"
                new_mask = pd.Series(True, index=df2.index)
                modified_mask = pd.Series(False, index=df2.index)  # ### нет модифицированных
            else:
                new_ids = curr_ids_set - prev_ids
                common_ids = curr_ids_set & prev_ids

                if prev_df is not None and common_ids:
                    common_columns = [
                        col for col in df2.columns
                        if col in prev_df.columns and col != id_col
                    ]

                    if common_columns:
                        common_ids_sorted = sorted(common_ids)

                        curr_subset = (
                            df2.set_index(id_col)
                            .loc[common_ids_sorted, common_columns]
                            .sort_index()
                        )
                        prev_subset = (
                            prev_df.set_index(id_col)
                            .loc[common_ids_sorted, common_columns]
                            .sort_index()
                        )

                        curr_norm = _normalize_for_compare(curr_subset)
                        prev_norm = _normalize_for_compare(prev_subset)

                        # ### В некоторых файлах встречаются дубли имён колонок,
                        # ### что может ломать прямое сравнение DataFrame.
                        # ### Предварительно выравниваем обе матрицы.
                        curr_norm, prev_norm = curr_norm.align(
                            prev_norm,
                            join='outer',
                            axis=None,
                            fill_value=''
                        )

                        diff_df = curr_norm != prev_norm  # ### поэлементные отличия
                        changed_series = diff_df.any(axis=1)
                        changed_ids = set(changed_series[changed_series].index)

                        # ### Сохраняем по id список изменённых колонок
                        for _id in changed_ids:
                            changed_cols = set(diff_df.columns[diff_df.loc[_id]])
                            if changed_cols:
                                changed_cells[_id] = changed_cols

                # NEW — только новые id
                new_mask = curr_ids.isin(new_ids)
                # MODIFIED — id, которые были и поменялись
                modified_mask = curr_ids.isin(changed_ids)

                # DELETED — только реально исчезнувшие
                deleted_ids_only = prev_ids - curr_ids_set

                if deleted_ids_only and prev_df is not None:
                    deleted_rows_df = prev_df[prev_df[id_col].isin(deleted_ids_only)].copy()

                    for column in df2.columns:
                        if column not in deleted_rows_df.columns:
                            deleted_rows_df[column] = ""

            # Собираем итоговый DataFrame для листа
            df_out = df2.copy()
            if add_flag_column:
                df_out.insert(0, "Новая запись?", "")

            # Статусы по строкам текущего месяца
            status_series = pd.Series("", index=df2.index, dtype=object)
            status_series.loc[new_mask] = "NEW"
            status_series.loc[modified_mask] = "MODIFIED"  # ### новые статусы

            if add_flag_column:
                df_out.loc[:, "Новая запись?"] = status_series

            # Добавляем удалённые строки (из предыдущего месяца)
            if not deleted_rows_df.empty:
                if add_flag_column:
                    deleted_rows_df.insert(0, "Новая запись?", "DELETED")

                # Синхронизируем набор и порядок колонок
                extra_cols_for_out = [
                    column for column in deleted_rows_df.columns if column not in df_out.columns
                ]
                for column in extra_cols_for_out:
                    df_out[column] = ""

                missing_in_deleted = [
                    column for column in df_out.columns if column not in deleted_rows_df.columns
                ]
                for column in missing_in_deleted:
                    deleted_rows_df[column] = ""

                deleted_rows_df = deleted_rows_df.loc[:, df_out.columns]
                df_out = pd.concat([df_out, deleted_rows_df], ignore_index=True, sort=False)

            # Формируем список статусов в порядке строк df_out
            statuses_list = status_series.tolist()
            if not deleted_rows_df.empty:
                statuses_list.extend(["DELETED"] * len(deleted_rows_df))

            if add_flag_column:
                df_out["Новая запись?"] = df_out["Новая запись?"].fillna("")

            # Пишем лист
            sheet_name = str(key)
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # --- ФОРМАТИРОВАНИЕ ПРОЦЕНТОВ И ФИНАНСОВ ---
            headers = [cell.value for cell in ws[1]]
            header_index = {name: idx + 1 for idx, name in enumerate(headers)}

            # 1) Процентные колонки
            percent_columns = [
                "% Завершенности проекта",
                "% валовая маржинальность",
                "% прибыли"
            ]
            percent_format = "0.0%"

            for col_name in percent_columns:
                if col_name in header_index:
                    col = header_index[col_name]
                    for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                        for c in cell:
                            c.number_format = percent_format

            # 2) Денежные колонки
            money_columns = [
                "Реализация без НДС",
                "Чистые продажи",
                "Прибыль",
                "Unnamed: 18",
                "Расходы подрядчиков",
                "Расходы CATI по проектам",
                "Заработная плата по проектам (ГПХ)",
                "Материалы, лицензии и аренда помещений по проектам",
                "Командировки по проектам",
                "Отправка посылок по проектам",
                "Межгород по проектам",
                "Итого прямые расходы",
                "Unnamed: 27",
                "Timesheets",
                "CATI/Онлайн панель overheads",
                "Корректировка",
                "Итого операционные расходы",
                # summary
                "Сумма 'Реализация без НДС'",
                "Сумма 'Total Direct Costs'",
                "Сумма 'Total Operating Costs'",
                "Operation Profit"
            ]

            rub_format = '# ##0'  # без копеек, с пробелами

            for col_name in money_columns:
                if col_name in header_index:
                    col = header_index[col_name]
                    col_letter = ws.cell(row=1, column=col).column_letter

                    # ширина столбца
                    ws.column_dimensions[col_letter].width = 18

                    # применяем формат
                    for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                        for c in cell:
                            c.number_format = rub_format

            # --- КОНЕЦ ФОРМАТИРОВАНИЕ ПРОЦЕНТОВ И ФИНАНСОВ ---

            ncols = df_out.shape[1]
            id_list = df2[id_col].tolist()  # ### id только для «текущих» строк

            # Подсветка строк / ячеек
            for row_idx, status in enumerate(statuses_list, start=2):
                # индекс относительно текущего месяца
                idx_in_current = row_idx - 2

                if status == "NEW":
                    # подсветка всей строки зелёным
                    for col_idx in range(1, ncols + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = green_fill

                elif status == "MODIFIED":
                    # ### подсвечиваем только изменённые ячейки жёлтым
                    if 0 <= idx_in_current < len(id_list):
                        row_id = id_list[idx_in_current]
                        changed_cols_for_id = changed_cells.get(row_id, set())
                        if changed_cols_for_id:
                            for col_idx in range(1, ncols + 1):
                                col_name = df_out.columns[col_idx - 1]
                                if col_name in changed_cols_for_id:
                                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

                elif status == "DELETED":
                    # подсветка всей строки красным
                    for col_idx in range(1, ncols + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = red_fill

                # остальные статусы (пустые) — без подсветки

            prev_ids = curr_ids_set
            prev_df = df2.copy()

    return output_path


def save_summary_with_format(df, output_path):
    # Сохраняем без форматирования сначала
    df.to_excel(output_path, index=False)

    # Загружаем файл через openpyxl
    wb = load_workbook(output_path)
    ws = wb.active

    # Заголовки
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx + 1 for idx, name in enumerate(headers)}

    # -----------------------------
    # 1) ПРОЦЕНТНЫЕ КОЛОНКИ xx.x%
    # -----------------------------
    percent_columns = [
        "YTD vs. Таргет",
    ]

    percent_format = "0.0%"  # один знак после запятой

    for col_name in percent_columns:
        if col_name in header_index:
            col = header_index[col_name]
            for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                for c in cell:
                    c.number_format = percent_format

    # ------------------------------------------------------------
    # 2) ФИНАНСОВЫЕ КОЛОНКИ (xxx xxx без десятичных, рубли)
    # ------------------------------------------------------------
    money_columns = [
        "Сумма 'Реализация без НДС'",
        "Сумма 'Total Direct Costs'",
        "Сумма 'Total Operating Costs'",
        "Operation Profit",
        "YTD",
        "Бюджет - таргет",
        "Q1RF",
        "Q2RF",
        "Q3RF",
        "X-charge",
    ]

    # между # и ##0 тут — НЕРАЗРЫВНЫЙ ПРОБЕЛ (U+00A0)
    rub_format = '# ##0'   # без копеек

    for col_name in money_columns:
        if col_name in header_index:
            col = header_index[col_name]
            col_letter = ws.cell(row=1, column=col).column_letter

            # Ширина столбца
            ws.column_dimensions[col_letter].width = 18

            # Форматирование значений
            for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                for c in cell:
                    c.number_format = rub_format

    wb.save(output_path)

